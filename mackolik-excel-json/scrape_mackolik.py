#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Mackolik.com'dan günlük maç verilerini scrape edip Excel'e yazma (Selenium ile)
"""
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from openpyxl import Workbook
from datetime import datetime
import time

def scrape_mackolik():
    """Mackolik.com'dan HTML table'ı Selenium ile scrape et ve Excel'e yaz"""
    
    url = "https://arsiv.mackolik.com/Genis-Iddaa-Programi"
    
    print(f"[INFO] Mackolik'ten veriler çekiliyor: {url}")
    print(f"[INFO] Tarih: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}\n")
    
    driver = None
    try:
        # Selenium Chrome driver
        options = webdriver.ChromeOptions()
        options.add_argument('--headless')  # GUI olmadan çalış
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        
        print("[INFO] Browser baslatiliyor...")
        driver = webdriver.Chrome(options=options)
        driver.get(url)
        
        print("[INFO] Sayfa yüklendi, JavaScript çalıştırılıyor...")
        
        # Eger form varsa, formu incele ve submit et
        try:
            form = driver.find_element(By.CSS_SELECTOR, "form[method='get']")
            print("[INFO] Form bulundu, submit ediliyor...")
            form.submit()
            time.sleep(3)  # Form submit sonrasi bekle
        except:
            print("[INFO] Form bulunamadi, sayfada devam ediliyor...")
        
        # Table'ın yüklenmesini bekle (max 15 saniye)
        wait = WebDriverWait(driver, 15)
        try:
            # Table'ın en az birkaç satır içermesini bekle
            table = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "table[width='100%'][class='table-header']")))
            print("[OK] Table bulundu")
            
            # JavaScript'in render etmesi için biraz bekle
            time.sleep(2)
        except:
            print("[WARN] Table loading timeout, yinede denemeye devam ediliyor...")
        
        # Page source'u al
        page_source = driver.page_source
        
        # DEBUG: HTML'ı dosyaya kaydet
        with open('mackolik_debug.html', 'w', encoding='utf-8') as f:
            f.write(page_source)
        print("[DEBUG] HTML kaydedildi: mackolik_debug.html")
        
        # HTML parse
        soup = BeautifulSoup(page_source, 'html.parser')
        
        # Table bul
        table = soup.find('table', {'width': '100%', 'class': 'table-header'})
        if not table:
            # Alternatif: ilk table'ı bul
            table = soup.find('table', {'class': 'table-header'})
        
        if not table:
            print("[ERROR] Table bulunamadi!")
            return False
        
        print(f"[OK] Table bulundu")
        
        # Tüm satırları oku
        rows = table.find_all('tr')
        print(f"[INFO] {len(rows)} satır bulundu")
        
        if len(rows) < 2:
            print("[ERROR] Yeterince veri yok!")
            return False
        
        # Excel workbook oluştur
        wb = Workbook()
        ws = wb.active
        
        # Header'ı yaz
        header_row = rows[0]
        header_cells = header_row.find_all('td')
        for col_idx, cell in enumerate(header_cells, 1):
            ws.cell(row=1, column=col_idx, value=cell.get_text(strip=True))
        
        print(f"[OK] Header yazildi ({len(header_cells)} sütun)")
        
        # Data satırlarını yaz
        data_count = 0
        for row_idx, row in enumerate(rows[1:], 2):
            cells = row.find_all('td')
            for col_idx, cell in enumerate(cells, 1):
                value = cell.get_text(strip=True)
                # Sayısal değerleri parse et
                try:
                    if '.' in value and len(value) < 10:
                        value = float(value)
                except:
                    pass
                ws.cell(row=row_idx, column=col_idx, value=value)
            
            if len(cells) > 0:
                data_count += 1
        
        print(f"[OK] {data_count} satir veri yazildi")
        
        # Dosya kaydet
        today = datetime.now().strftime('%d.%m.%Y')
        filename = 'gunlukmaclar.xlsx'
        filepath = filename
        
        wb.save(filepath)
        print(f"[OK] Dosya kaydedildi: {filepath}")
        print(f"[SUCCESS] Toplam {data_count} mac basarili sekilde indirildi!\n")
        
        return True
        
    except Exception as e:
        print(f"[ERROR] Hata: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    finally:
        if driver:
            driver.quit()
            print("[INFO] Browser kapatildi")

if __name__ == "__main__":
    success = scrape_mackolik()
    exit(0 if success else 1)
