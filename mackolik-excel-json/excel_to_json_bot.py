#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Mackolik Ham Excel (Y) → Sadeleştirilmiş Excel (X) + JSON dönüştürücü

Y Excel sütunları:
  1: saat, 2: lig, 6: ev sahibi, 7: misafir
  10: MS 1, 11: MS X, 12: MS 2
  21: KG Var, 22: KG Yok
  28: 2.5 Alt, 29: 2.5 Üst
  30: 3.5 Alt, 31: 3.5 Üst
  32: 4.5 Alt, 33: 4.5 Üst
  Satır 3: tarih, Satır 4+: maç verileri
"""
import openpyxl
from openpyxl import Workbook
import json
import os
import sys
from datetime import datetime

# Windows encoding fix
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.stderr.reconfigure(encoding='utf-8', errors='replace')


def safe_float(val):
    """Değeri float'a çevir, başarısızsa 0 döndür."""
    if val is None or val == '-' or val == '' or val == 0:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def find_input_file():
    """mackolik-excel-json klasöründe gunlukmaclar.xlsx dosyasını bul."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    target = os.path.join(script_dir, 'gunlukmaclar.xlsx')
    if os.path.exists(target):
        return target

    # Alternatif: klasördeki herhangi bir xlsx (json_output hariç)
    for f in os.listdir(script_dir):
        if f.endswith('.xlsx') and f != 'gunlukmaclar.xlsx':
            full = os.path.join(script_dir, f)
            if os.path.isfile(full):
                return full

    return None


def convert(input_file, output_dir):
    """Y Excel → X Excel + JSON dönüşümü yapar."""
    print(f"📂 Okuyor: {input_file}")

    wb_kaynak = openpyxl.load_workbook(input_file)
    ws_kaynak = wb_kaynak.active

    print(f"✓ {ws_kaynak.max_row} satır, {ws_kaynak.max_column} sütun")

    # Tarih bilgisini al (satır 3, sütun 1)
    raw_date = ws_kaynak.cell(3, 1).value
    if isinstance(raw_date, datetime):
        date_str = raw_date.strftime('%d.%m.%Y')
    elif raw_date:
        date_str = str(raw_date).split(' ')[0]
        # YYYY-MM-DD formatını DD.MM.YYYY'ye çevir
        try:
            dt = datetime.strptime(date_str, '%Y-%m-%d')
            date_str = dt.strftime('%d.%m.%Y')
        except ValueError:
            pass
    else:
        date_str = datetime.now().strftime('%d.%m.%Y')

    print(f"📅 Tarih: {date_str}")

    # output klasörünü oluştur
    os.makedirs(output_dir, exist_ok=True)

    # ===== X Excel oluştur =====
    wb_yeni = Workbook()
    ws_yeni = wb_yeni.active

    # Başlık satırı 1
    ws_yeni['A1'] = 'Home'
    ws_yeni['B1'] = 'Away'
    ws_yeni['C1'] = 1
    ws_yeni['D1'] = 0
    ws_yeni['E1'] = 2
    ws_yeni['F1'] = 'Karsılıklı Gol'
    ws_yeni['H1'] = '2.5 üst'
    ws_yeni['J1'] = '3.5 üst'
    ws_yeni['L1'] = '5.5üst'

    # Başlık satırı 2
    ws_yeni['A2'] = 'Ev Sahibi'
    ws_yeni['B2'] = 'Misafir'
    ws_yeni['C2'] = 1.00
    ws_yeni['D2'] = 0.00
    ws_yeni['E2'] = 2.00
    ws_yeni['F2'] = 'Var'
    ws_yeni['G2'] = 'Yok'
    ws_yeni['H2'] = 'Alt'
    ws_yeni['I2'] = 'Üst'
    ws_yeni['J2'] = 'Alt'
    ws_yeni['K2'] = 'Üst'
    ws_yeni['L2'] = '5.5üst'

    # ===== JSON verisi =====
    matches_json = []

    # Veriyi dönüştür (satır 4'ten başla, 3 tarih satırı)
    yeni_satir = 3
    for satir in range(4, ws_kaynak.max_row + 1):
        ev_sahibi = ws_kaynak.cell(satir, 6).value   # F: Ev Sahibi
        misafir = ws_kaynak.cell(satir, 7).value      # G: Misafir

        if not ev_sahibi or not misafir:
            continue

        ev_sahibi = str(ev_sahibi).strip()
        misafir = str(misafir).strip()

        if not ev_sahibi or not misafir:
            continue

        # Saat bilgisi
        saat_raw = ws_kaynak.cell(satir, 1).value
        if hasattr(saat_raw, 'strftime'):
            saat = saat_raw.strftime('%H:%M')
        elif saat_raw:
            saat = str(saat_raw).split(':')[0] + ':' + str(saat_raw).split(':')[1] if ':' in str(saat_raw) else str(saat_raw)
        else:
            saat = ''

        # Oranları oku
        ms_1 = safe_float(ws_kaynak.cell(satir, 10).value)    # J: MS 1
        ms_x = safe_float(ws_kaynak.cell(satir, 11).value)    # K: MS X
        ms_2 = safe_float(ws_kaynak.cell(satir, 12).value)    # L: MS 2
        kg_var = safe_float(ws_kaynak.cell(satir, 21).value)   # U: KG Var
        kg_yok = safe_float(ws_kaynak.cell(satir, 22).value)   # V: KG Yok
        alt_2_5 = safe_float(ws_kaynak.cell(satir, 28).value)  # 2.5 Alt
        ust_2_5 = safe_float(ws_kaynak.cell(satir, 29).value)  # 2.5 Üst
        alt_3_5 = safe_float(ws_kaynak.cell(satir, 30).value)  # 3.5 Alt
        ust_3_5 = safe_float(ws_kaynak.cell(satir, 31).value)  # 3.5 Üst
        ust_5_5 = safe_float(ws_kaynak.cell(satir, 33).value)  # 4.5 Üst (en yakın 5.5'e)

        # X Excel'e yaz
        ws_yeni.cell(yeni_satir, 1, ev_sahibi)    # A: Home
        ws_yeni.cell(yeni_satir, 2, misafir)       # B: Away
        ws_yeni.cell(yeni_satir, 3, ms_1)          # C: 1
        ws_yeni.cell(yeni_satir, 4, ms_x)          # D: X
        ws_yeni.cell(yeni_satir, 5, ms_2)          # E: 2
        ws_yeni.cell(yeni_satir, 6, kg_var)        # F: KG Var
        ws_yeni.cell(yeni_satir, 7, kg_yok)        # G: KG Yok
        ws_yeni.cell(yeni_satir, 8, alt_2_5)       # H: 2.5 Alt
        ws_yeni.cell(yeni_satir, 9, ust_2_5)       # I: 2.5 Üst
        ws_yeni.cell(yeni_satir, 10, alt_3_5)      # J: 3.5 Alt
        ws_yeni.cell(yeni_satir, 11, ust_3_5)      # K: 3.5 Üst
        ws_yeni.cell(yeni_satir, 12, ust_5_5)      # L: 5.5 Üst

        yeni_satir += 1

        # JSON'a ekle
        matches_json.append({
            'home_team': ev_sahibi,
            'away_team': misafir,
            'saat': saat,
            'ms_1': ms_1,
            'ms_x': ms_x,
            'ms_2': ms_2,
            'kg_var': kg_var,
            'kg_yok': kg_yok,
            'ust_2_5': ust_2_5,
            'alt_2_5': alt_2_5,
            'ust_3_5': ust_3_5,
            'alt_3_5': alt_3_5,
            'ust_5_5': ust_5_5
        })

    mac_sayisi = yeni_satir - 3

    # X Excel kaydet
    excel_output = os.path.join(output_dir, f"{date_str.replace('.', '')}.xlsx")
    wb_yeni.save(excel_output)
    print(f"📊 Excel kaydedildi: {excel_output} ({mac_sayisi} maç)")

    # JSON kaydet
    json_data = {
        'date': date_str,
        'total_matches': mac_sayisi,
        'matches': matches_json
    }

    json_output = os.path.join(output_dir, f"{date_str.replace('.', '')}.json")
    with open(json_output, 'w', encoding='utf-8') as f:
        json.dump(json_data, f, ensure_ascii=False, indent=2)

    print(f"📋 JSON kaydedildi: {json_output} ({mac_sayisi} maç)")
    print(f"✅ {mac_sayisi} maç başarıyla işlendi!")

    return True


def main():
    print("\n" + "=" * 60)
    print("  📊 MACKOLIK EXCEL → JSON DÖNÜŞTÜRÜCÜ")
    print("=" * 60)

    # Input dosyasını bul
    if len(sys.argv) >= 2:
        input_file = sys.argv[1]
    else:
        input_file = find_input_file()

    if not input_file or not os.path.exists(input_file):
        print("❌ gunlukmaclar.xlsx bulunamadı!")
        print("   Lütfen dosyayı mackolik-excel-json klasörüne koyun.")
        sys.exit(1)

    # Output klasörü
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(script_dir, 'json_output')

    success = convert(input_file, output_dir)

    if not success:
        sys.exit(1)


if __name__ == "__main__":
    main()
